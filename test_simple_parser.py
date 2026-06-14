#!/usr/bin/env python3
"""
Simple test for the DocumentParser functionality.
"""

import tempfile
import os
from pathlib import Path

# Add the project to the path
project_path = Path(__file__).parent
import sys
sys.path.insert(0, str(project_path))

from timetable_system.document_parser import DocumentParser


def create_test_excel_file():
    """Create a simple test Excel file."""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Courses"

        # Simple headers
        headers = ["Course Code", "Course Name", "Lecturer"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Simple test data
        test_data = [
            ["CS101", "Introduction to CS", "Mr F. Dube"],
            ["CS102", "Data Structures", "Dr Sakala"],
            ["CS201", "Algorithms", "TBA"],
        ]

        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save to temp file
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = tmp.name
        tmp.close()

        wb.save(tmp_path)
        wb.close()

        return tmp_path

    except ImportError as e:
        print(f"Error importing openpyxl: {e}")
        return None


def test_parser():
    """Test the document parser."""
    print("Testing DocumentParser...")

    # Create test file
    test_file = create_test_excel_file()

    if not test_file:
        print("Failed to create test Excel file")
        return

    try:
        # Parse the file
        parser = DocumentParser()
        parsed_data = parser.parse_all_formats(test_file)

        print("\n=== PARSED DATA ===")
        print(f"Source file: {parsed_data.get('source_file', 'N/A')}")
        print(f"Notes: {parsed_data.get('notes', 'N/A')}")
        print(f"\nCourses: {len(parsed_data.get('courses', []))}")
        for course in parsed_data.get('courses', []):
            print(f"  - {course.get('course_code', '')}: {course.get('course_name', '')}")
            print(f"    Lecturer: {course.get('lecturer_username', 'None')}")

        print(f"\nLecturers: {len(parsed_data.get('lecturers', []))}")
        for lecturer in parsed_data.get('lecturers', []):
            print(f"  - {lecturer.get('full_name', '')}: {lecturer.get('username', '')} <{lecturer.get('email', '')}>")

        print(f"\nRooms: {len(parsed_data.get('rooms', []))}")

        # Export to JSON
        json_output = parser.export_to_json(parsed_data)
        import json
        parsed_json = json.loads(json_output)

        print(f"\n=== JSON EXPORT ===")
        print(f"Courses in JSON: {len(parsed_json.get('courses', []))}")
        print(f"Lecturers in JSON: {len(parsed_json.get('lecturers', []))}")
        print(f"Rooms in JSON: {len(parsed_json.get('rooms', []))}")

        # Check for required fields
        print("\n=== FIELD VALIDATION ===")
        for course in parsed_json.get('courses', []):
            required_fields = ['course_code', 'course_name', 'duration_hours']
            for field in required_fields:
                if field not in course:
                    print(f"  Missing field '{field}' in course {course.get('course_code', 'UNKNOWN')}")
                else:
                    print(f"  Course {course.get('course_code', '')}: {field} = {course[field]}")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

    finally:
        # Clean up
        if 'test_file' in locals() and os.path.exists(test_file):
            os.unlink(test_file)
            print(f"\nCleaned up temp file: {test_file}")


if __name__ == '__main__':
    test_parser()