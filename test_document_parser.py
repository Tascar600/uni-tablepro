#!/usr/bin/env python3
"""
Test script for the DocumentParser functionality.
"""

import os
import tempfile
import json
from pathlib import Path

# Add the project to the path
project_path = Path(__file__).parent / "timetable_system"
import sys
sys.path.insert(0, str(project_path))

from document_parser import DocumentParser


def create_test_excel_file():
    """Create a test Excel file with various data structures."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Courses"

    # Write header row with flexible column names
    headers = ["Course Code", "Course Name", "Lecturer", "Hours", "Max Students", "Level", "Group"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Add test data
    test_data = [
        ["CS101", "Introduction to Computer Science", "Mr F. Dube", 4, 50, "department", "1.1"],
        ["CS102", "Data Structures", "Dr Sakala", 4, 45, "university", "1.2"],
        ["CS201", "Algorithms", "TBA", 3, 40, "department", "2.1"],
        ["CS301", "Software Engineering", "Mr Mhlanganiso", 4, 35, "department", "3.1"],
        ["MATH101", "Calculus I", "", 3, 60, "university", "1.1"],
    ]

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Create a second sheet with different structure
    ws2 = wb.create_sheet("Lecturers")
    headers2 = ["Name", "Title", "Email", "Department"]
    for col_idx, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col_idx, value=header)

    lecturers = [
        ["Dr Alice M. Banda", "Professor", "abanda@buse.ac.zw", "Mathematics"],
        ["Mr B. Chikwanda", "Lecturer", "bchikwanda@buse.ac.zw", "Computer Science"],
        ["Ms C. Ndlovu", "Assistant Lecturer", "cndlovu@buse.ac.zw", "Mathematics"],
    ]

    for row_idx, row_data in enumerate(lecturers, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=cell_value)

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()

    wb.save(tmp_path)
    wb.close()

    return tmp_path


def test_excel_parsing():
    """Test Excel parsing functionality."""
    print("Testing Excel parsing...")

    # Create test file
    test_file = create_test_excel_file()

    # Parse the file
    parser = DocumentParser()
    parsed_data = parser.parse_all_formats(test_file)

    # Print results
    print("\n=== PARSED DATA ===")
    print(f"Source file: {parsed_data.get('source_file', 'N/A')}")
    print(f"Notes: {parsed_data.get('notes', 'N/A')}")
    print(f"\nCourses: {len(parsed_data.get('courses', []))}")
    for course in parsed_data.get('courses', []):
        print(f"  - {course.get('course_code', '')}: {course.get('course_name', '')}")
        print(f"    Lecturer: {course.get('lecturer_username', 'None')}, Group: {course.get('group', 'None')}")

    print(f"\nLecturers: {len(parsed_data.get('lecturers', []))}")
    for lecturer in parsed_data.get('lecturers', []):
        print(f"  - {lecturer.get('full_name', '')}: {lecturer.get('username', '')} <{lecturer.get('email', '')}>")

    print(f"\nRooms: {len(parsed_data.get('rooms', []))}")

    # Export to JSON
    json_output = parser.export_to_json(parsed_data)
    parsed_json = json.loads(json_output)

    print(f"\n=== JSON EXPORT ===")
    print(f"Courses in JSON: {len(parsed_json.get('courses', []))}")
    print(f"Lecturers in JSON: {len(parsed_json.get('lecturers', []))}")
    print(f"Rooms in JSON: {len(parsed_json.get('rooms', []))}")\n    # Verify the schema
    required_fields = {
        'course_code': '', 'course_name': '', 'duration_hours': 4,
        'level': '', 'department': '', 'color': 'blue',
        'max_students': 50, 'group': None, 'lecturer_username': None,
        'lecturer_email': None, 'notes': None, 'source_location': None, 'flag': None
    }

    print("\n=== SCHEMA VALIDATION ===")
    for course in parsed_json.get('courses', []):
        for field, default in required_fields.items():
            if field not in course:
                print(f"  WARNING: Missing field '{field}'")
            elif course[field] is None and default is not None:
                print(f"  WARNING: Field '{field}' is null (expected default: {default})")
            elif course[field] != default and default is not None:
                print(f"  INFO: Field '{field}' = {course[field]} (default: {default})")

    # Clean up
    os.unlink(test_file)

    return parsed_json


def main():
    """Main test function."""
    print("=== Document Parser Test ===\n")

    # Test Excel parsing
    test_excel_parsing()

    print("\n=== Test completed successfully ===")


if __name__ == '__main__':
    main()